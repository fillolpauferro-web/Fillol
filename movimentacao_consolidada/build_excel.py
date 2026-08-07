"""Grava o resultado do ETL num .xlsx com abas prontas para usar no Excel.

- "Painel Cliente Faturado" / "Painel Cliente Reservado": escolhe o cliente
  num dropdown e vê o waterfall daquele painel (fórmulas SUMIFS em cima da
  tabela pequena -- instantâneo). Cada painel tem seu próprio Saldo
  Inicial/Final, definidos em config.PAINEIS.
- "Geral": waterfall consolidado (todos os clientes) do painel principal
  (config.PAINEL_PRINCIPAL).
- "Por Cliente": tabela fato do painel principal (uma linha por cliente x
  mês) que alimenta o Painel Cliente Faturado / Tabela Dinâmica / Power
  Query, mais uma tabela auxiliar mais abaixo que alimenta o Painel Cliente
  Reservado.
- "Notas Fiscais": detalhe (não agregado) por nota, só dos Document Type em
  config.NOTAS_FISCAIS_TIPOS_DOCUMENTO.
- "Fato (base Python)": granularidade cliente x mês x categoria, caso
  precise auditar.
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import config

NUM_FMT = "#,##0"


def _fmt_mes(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series).dt.strftime("%b/%y")


def _fmt_por_cliente(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["mes"] = _fmt_mes(df["mes"])
    return df.rename(columns={"cnpj_raiz": "CNPJ", "cliente": "Cliente", "mes": "Data"})


def _fmt_geral(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["mes"] = _fmt_mes(df["mes"])
    return df.rename(columns={"mes": "Data"})


def _write_table(ws, df: pd.DataFrame, table_name: str, valor_cols: set, start_row: int = 1) -> tuple:
    """Escreve df como Excel Table a partir de start_row (coluna A).
    Retorna (última linha de dados, próxima linha livre depois da tabela)."""
    for j, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col_name)
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row, start=1):
            # NaN (célula vazia do pandas) grava errado no xlsx e pode gerar
            # "arquivo corrompido" no Excel -- vira célula em branco.
            if isinstance(val, float) and pd.isna(val):
                val = None
            ws.cell(row=start_row + i, column=j, value=val)

    n_rows, n_cols = df.shape
    last_col = get_column_letter(n_cols)
    last_row = start_row + n_rows
    table = Table(displayName=table_name, ref=f"A{start_row}:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col_name)) + 2)
        if col_name in valor_cols:
            for r in range(start_row + 1, last_row + 1):
                ws.cell(row=r, column=col_idx).number_format = NUM_FMT

    if start_row == 1:
        ws.freeze_panes = "A2"
    return last_row, last_row + 2  # deixa 1 linha de folga antes da próxima tabela


def _sref(table_name: str, col_name: str) -> str:
    # Nome de coluna com caracteres especiais (ex.: "Canc. / Devolução")
    # precisa ficar entre colchetes na referência estruturada do Excel.
    return f"{table_name}[{col_name}]"


def _write_painel_sheet(
    ws, por_cliente_painel: pd.DataFrame, categorias: list, tabela_fonte: str, nome_lista_clientes: str
) -> None:
    ws["A1"] = "Cliente:"
    clientes_ordenados = sorted(por_cliente_painel["Cliente"].dropna().unique().tolist())
    ws["B1"] = clientes_ordenados[0] if clientes_ordenados else ""
    ws["A1"].font = ws["B1"].font.copy(bold=True)

    # A fonte da lista precisa ser um Intervalo Nomeado (não uma referência
    # estruturada tipo Tabela[Coluna]) porque a validação de dados fica numa
    # aba diferente de onde a tabela vive -- referência estruturada
    # entre abas em Data Validation corrompe o arquivo em alguns Excel.
    dv = DataValidation(type="list", formula1=f"={nome_lista_clientes}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])

    meses = por_cliente_painel[["Data"]].drop_duplicates()
    meses["_ord"] = pd.to_datetime(meses["Data"], format="%b/%y")
    meses = meses.sort_values("_ord")["Data"].tolist()

    header_row = 3
    headers = ["Data", "Saldo Inicial", *categorias, "Saldo Final"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)

    for k, mes in enumerate(meses):
        r = header_row + 1 + k
        ws.cell(row=r, column=1, value=mes)
        for j, col_name in enumerate(headers[1:], start=2):
            formula = (
                f"=SUMIFS({_sref(tabela_fonte, col_name)},"
                f"{_sref(tabela_fonte, 'Cliente')},$B$1,"
                f"{_sref(tabela_fonte, 'Data')},$A{r})"
            )
            cell = ws.cell(row=r, column=j, value=formula)
            cell.number_format = NUM_FMT

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.freeze_panes = "A4"


def write_workbook(
    fato: pd.DataFrame,
    geral_por_painel: dict,
    por_cliente_por_painel: dict,
    paineis: list,
    categorias_fato: list,
    notas_fiscais: pd.DataFrame,
    excel_path: Path = None,
) -> None:
    excel_path = Path(excel_path) if excel_path else config.EXCEL_OUTPUT

    paineis_por_chave = {p["chave"]: p for p in paineis}
    principal = paineis_por_chave[config.PAINEL_PRINCIPAL]

    fato_fmt = fato.copy()
    fato_fmt["mes"] = _fmt_mes(fato_fmt["mes"])
    fato_fmt = fato_fmt.rename(columns={
        "cnpj_raiz": "CNPJ", "cliente": "Cliente", "mes": "Data",
        "categoria": "Categoria", "valor": "Valor",
    })

    notas_fmt = notas_fiscais.copy()
    notas_fmt["mes"] = _fmt_mes(notas_fmt["mes"])
    notas_fmt = notas_fmt.rename(columns={
        "mes": "Mês", "cliente": "Nome Cliente", "cnpj_raiz": "CNPJ",
        "nfe_number": "NFE Number", "nfe_item": "NFE Item", "po_number": "PO Number",
        "valor_confirmado": "Value Confirmed",
    })

    wb = Workbook()
    wb.remove(wb.active)

    # --- Por Cliente: tabela principal (painel principal) + tabela(s) auxiliares ---
    ws_cli = wb.create_sheet("Por Cliente")
    tabelas_fonte = {}  # chave do painel -> nome da Excel Table que o alimenta
    nomes_lista_clientes = {}  # chave do painel -> nome do Intervalo Nomeado (coluna Cliente)
    proxima_linha = 1
    for painel in paineis:
        df_pc = _fmt_por_cliente(por_cliente_por_painel[painel["chave"]])
        valor_cols = {"Saldo Inicial", "Saldo Final", *painel["categorias"]}
        nome_tabela = "TabelaPorCliente" if painel["chave"] == principal["chave"] else f"TabelaPorCliente_{painel['chave']}"
        if painel["chave"] != principal["chave"]:
            ws_cli.cell(row=proxima_linha, column=1, value=f"Base para o painel \"{painel['aba']}\" (uso interno das fórmulas):")
            proxima_linha += 1
        start_row = proxima_linha
        last_row, proxima_linha = _write_table(ws_cli, df_pc, nome_tabela, valor_cols, start_row=start_row)
        tabelas_fonte[painel["chave"]] = nome_tabela

        nome_lista = f"ListaClientes_{painel['chave']}"
        ref = f"'Por Cliente'!$B${start_row + 1}:$B${last_row}"
        wb.defined_names[nome_lista] = DefinedName(nome_lista, attr_text=ref)
        nomes_lista_clientes[painel["chave"]] = nome_lista

    # --- Geral: painel principal ---
    ws_geral = wb.create_sheet("Geral")
    df_geral = _fmt_geral(geral_por_painel[principal["chave"]])
    valor_cols_geral = {"Saldo Inicial", "Saldo Final", *principal["categorias"]}
    _write_table(ws_geral, df_geral, "TabelaGeral", valor_cols_geral)

    # --- Notas Fiscais ---
    ws_notas = wb.create_sheet("Notas Fiscais")
    _write_table(ws_notas, notas_fmt, "TabelaNotasFiscais", {"Value Confirmed"})

    # --- Fato (base Python) ---
    ws_fato = wb.create_sheet("Fato (base Python)")
    _write_table(ws_fato, fato_fmt, "TabelaFato", {"Valor", *categorias_fato})

    # --- Um Painel Cliente por painel, cada um na sua própria aba ---
    for painel in reversed(paineis):  # reversed: create_sheet(..., 0) empilha, então o 1º da lista fica em 1º
        ws_painel = wb.create_sheet(painel["aba"], 0)
        df_pc = _fmt_por_cliente(por_cliente_por_painel[painel["chave"]])
        _write_painel_sheet(
            ws_painel, df_pc, painel["categorias"],
            tabelas_fonte[painel["chave"]], nomes_lista_clientes[painel["chave"]],
        )

    try:
        wb.save(excel_path)
    except PermissionError:
        raise PermissionError(
            f"Não consegui salvar {excel_path.name} porque o arquivo está aberto "
            "no Excel (ou outro programa). Feche o Excel e rode de novo."
        ) from None
